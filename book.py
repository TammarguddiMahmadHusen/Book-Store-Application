# class Book:
#     def __init__(self, name, author, id, price):
#         self.name = name
#         self.author = author
#         self.id = id
#         self.price = price

#     def __str__(self):
#         return f"ID: {self.id}\tName: {self.name}\tAuthor: {self.author}\tPrice: {self.price}"

# books = []

# def search_id(id):
#     for book in books:
#         if book.id == id:
#             print(book)
#             return True
#     return False

# def search_name(name):
#     for book in books:
#         if book.name.lower() == name.lower():
#             print(book)
#             return True
#     return False

# def search_author(author):
#     for book in books:
#         if book.author.lower() == author.lower():
#             print(book)
#             return True
#     return False

# def list_books():
#     if books:
#         for book in books:
#             print(book)
#         return True
#     else:
#         return False

# def add_book(name, author, id, price):
#     for book in books:
#         if book.name.lower() == name.lower() or book.id == id:
#             return False
#     new_book = Book(name, author, id, price)
#     books.append(new_book)
#     return True

# def choice():
#     print("Enter the choice")
#     print("1: enter the new book details")
#     print("2: Search the book")
#     print("3: List of books")
#     print("4: exit")

# print("Welcome")

# while True:
#     choice()
#     ch = int(input())
#     if ch == 1:
#         print("Enter the name")
#         name = input()
#         print("Enter the author")
#         author = input()
#         print("Enter the id")
#         id = int(input())
#         print("Enter the price")
#         price = int(input())
#         if not add_book(name, author, id, price):
#             print("Name or id already exists")
#     elif ch == 2:
#         print("Enter 1: to search on id")
#         print("Enter 2: to search on name")
#         print("Enter 3: to search on author name")
#         c = int(input())
#         if c == 1:
#             print("Enter the id")
#             id = int(input())
#             if not search_id(id):
#                 print("Invalid input")
#         elif c == 2:
#             print("Enter the name")
#             name = input()
#             if not search_name(name):
#                 print("Invalid input")
#         elif c == 3:
#             print("Enter the author")
#             author = input()
#             if not search_author(author):
#                 print("Invalid input")
#         else:
#             print("Invalid input")
#     elif ch == 3:
#         list_books()
#     elif ch == 4:
#         break
#     else:
#         print("Invalid choice")

# # Importing library
# import qrcode

# # Data to be encoded
# data = 'QR Code using make() function'

# # Encoding data using make() function
# img = qrcode.make(data)

# # Saving as an image file
# img.save('MyQRCode1.png')


import turtle
import qrcode
from openpyxl import Workbook, load_workbook

# Initialize the Turtle screen
screen = turtle.Screen()
screen.setup(800, 400)
screen.title("Book Management System")

# Create a workbook for book data
workbook = Workbook()
sheet = workbook.active
sheet.append(["Book ID", "Title", "Author", "Price"])

# Define functions for the book system
def add_book():
    global sheet
    book_id = get_unique_book_id()
    title = input("Enter the title of the book: ")
    author = input("Enter the author of the book: ")
    price = input("Enter the price of the book: ")

    sheet.append([book_id, title, author, price])
    print("Book added successfully!")

def search_book(title):
    global sheet
    for row in sheet.iter_rows(values_only=True):
        if title.lower() in row[1].lower():
            print("Book ID:", row[0])
            print("Title:", row[1])
            print("Author:", row[2])
            print("Price:", row[3])
            return
    print("Book not found.")

def get_unique_book_id():
    global sheet
    if sheet.max_row == 1:
        return "B001"
    else:
        last_id = sheet.cell(row=sheet.max_row, column=1).value
        new_id = int(last_id[1:]) + 1
        return f"B{str(new_id).zfill(3)}"

def price_book(book_id):
    global sheet
    for row in sheet.iter_rows(values_only=True):
        if book_id == row[0]:
            price = row[3]
            print(f"The price of the book is ${price}")
            return
    print("Book ID not found.")

def generate_qr_code(book_id):
    for row in sheet.iter_rows(values_only=True):
        if book_id == row[0]:
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(f"Book ID: {row[0]}\nTitle: {row[1]}\nAuthor: {row[2]}\nPrice: {row[3]}")
            qr.make(fit=True)

            img = qr.make_image(fill_color="black", back_color="white")
            img.save(f"qrcodes/{row[0]}.png")
            print(f"QR Code for Book ID {row[0]} generated.")
            return
    print("Book ID not found.")

def save_to_excel():
    global workbook
    workbook.save("Book Store/books.xlsx")

# Main menu
while True:
    print("\nMain Menu:")
    print("1. Add a Book")
    print("2. Search for a Book")
    print("3. Price a Book")
    print("4. Generate QR Code")
    print("5. Save to Excel")
    print("6. Quit")
    choice = input("Enter your choice: ")

    if choice == "1":
        add_book()
    elif choice == "2":
        title = input("Enter the title of the book you want to search for: ")
        search_book(title)
    elif choice == "3":
        book_id = input("Enter the book ID to get the price: ")
        price_book(book_id)
    elif choice == "4":
        book_id = input("Enter the book ID to generate a QR code: ")
        generate_qr_code(book_id)
    elif choice == "5":
        save_to_excel()
        print("Data saved to Excel.")
    elif choice == "6":
        print("Exiting the program.")
        break

